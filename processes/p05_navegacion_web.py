"""
Módulo: processes.p05_navegacion_web
Descripción: Quinto y último subproceso del pipeline. Lee ``consolidado.xlsx``
             fila por fila, verifica en la plataforma transaccional de Salud Total
             EPS cada autorización cuyo ESTADO_ROBOT (columna G) esté vacío, y
             escribe el resultado —junto con 5 campos adicionales de la web— de
             vuelta al Excel en tiempo real (columnas G-L).
             Este subproceso siempre se ejecuta, aunque no haya PDFs nuevos,
             lo que permite reprocesar filas si el cliente borra manualmente el estado.
             Usa Playwright (Chromium) sin necesidad de ChromeDriver.
Autor: Brayand Javier Gomez Plata
"""

import os
import re
from typing import Any, Tuple

import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout


# ─── Selectores centralizados (REGLA 7) ───────────────────────────────────────
SELECTORES = {
    "tab_ips":                  'IPS',
    "tipo_doc_ips":             "TIPO DE DOCUMENTO DE LA IPS",
    "nit_ips":                  "NÚMERO DE IDENTIFICACIÓN DE",
    "tipo_doc_usuario":         "TIPO DE DOCUMENTO DEL USUARIO",
    "id_usuario":               "NÚMERO DEL USUARIO",
    "password":                 "CONTRASEÑA",
    "btn_ingresar":             "INGRESAR",
    "btn_direccionamientos":    "DIRECCIONAMIENTOS",
    "registrar_direccionamiento": "REGISTRAR DIRECCIONAMIENTO",
    "ventana_sede":             "DATOS SEDESUCURSAL SEDE",
    "combobox_sucursal":        "SELECCIONE UNA SUCURSAL",
    "combobox_sede":            "SELECCIONE UNA SEDE",
    "btn_aceptar":              "ACEPTAR",
    "campo_id_paciente":        "DIGITE NÚMERO DE IDENTIFICACI",
    "btn_consultar":            "CONSULTAR",
    "campo_nombre":             "DIGITE NOMBRE DEL PROTEGIDO",
    "consultar_dir":            "CONSULTAR DIRECCIONAMIENTO",
    "filtro_aut":               "DIRECCIONAMIENTO Filter",
}

# ─── Estados del robot ────────────────────────────────────────────────────────
ESTADO_ENCONTRADO    = "Registro encontrado"
ESTADO_NO_ENCONTRADO = "Examen NO se encuentra en Salud Total"
ESTADO_PAC_NO_EXISTE = "Paciente NO existe en el sistema"
ESTADO_NOMBRE_NO_OK  = "Nombre del paciente NO coincide"
ESTADO_ERROR         = "ERROR - MAX REINTENTOS"

# Columnas en consolidado.xlsx (1-based para openpyxl)
COL_ID_PACIENTE      = 3   # C
COL_NOMBRE           = 4   # D
COL_AUTORIZACION     = 5   # E
COL_COD_FACT         = 6   # F
COL_ESTADO_ROBOT     = 7   # G
COL_PAQUETE          = 8   # H
COL_CANTIDAD         = 9   # I
COL_SEDE             = 10  # J
COL_NOMBRE_CONVENIO  = 11  # K
COL_FECHA_VENCIMIENTO= 12  # L


# ─── Helpers de navegación ────────────────────────────────────────────────────

def _seleccionar_combobox(page, nombre: str, opcion: str, contenedor=None) -> None:
    """Selecciona una opción en un combobox Kendo UI de la plataforma SaludTotal.

    Hace clic en el campo, escribe el texto de la opción con un delay de 30 ms
    para simular tipeo humano, espera la opción en el dropdown y la selecciona.
    Si el dropdown no aparece en 2500 ms usa teclas de flecha como fallback.
    Verifica que el valor final coincida con la opción esperada.

    Args:
        page: Instancia de ``playwright.sync_api.Page``.
        nombre: Atributo ``name`` (o aria-label) del combobox.
        opcion: Texto exacto de la opción a seleccionar.
        contenedor: Locator opcional del elemento padre para acotar la búsqueda
                    (útil cuando hay múltiples comboboxes con el mismo nombre).

    Raises:
        RuntimeError: Si el valor final del combobox no contiene la opción esperada.
    """
    base = contenedor if contenedor else page
    cb = base.get_by_role("combobox", name=nombre)
    cb.wait_for(state="visible")
    cb.click()
    cb.fill("")
    cb.type(opcion, delay=30)

    opcion_lista = page.get_by_role("option", name=opcion, exact=True).first
    try:
        opcion_lista.wait_for(state="visible", timeout=2500)
        opcion_lista.click()
    except PlaywrightTimeout:
        cb.press("ArrowDown")
        cb.press("Enter")

    valor = cb.input_value().strip().upper()
    if opcion.strip().upper() not in valor:
        raise RuntimeError(
            f"No se pudo seleccionar '{opcion}' en '{nombre}'. Valor actual: '{valor}'"
        )


def _login(page, config: dict) -> None:
    """Realiza el login en la plataforma transaccional de SaludTotal EPS.

    Navega a la URL configurada, selecciona la pestaña IPS, rellena los
    campos de tipo de documento e identificación de la IPS y del usuario,
    ingresa la contraseña desde la variable de entorno ``PASSWORD_USUARIO``
    (nunca en código) y hace clic en INGRESAR.

    Args:
        page: Instancia de ``playwright.sync_api.Page`` con sesión activa.
        config: Configuración completa del robot. Se usan las claves
                ``url_plataforma``, ``tipo_doc_ips``, ``nit_ips``,
                ``tipo_doc_usuario`` e ``id_usuario``.
    """
    password = os.getenv("PASSWORD_USUARIO", "")
    tiempos = config.get("tiempos", {})

    page.goto(config["url_plataforma"])
    page.wait_for_load_state("networkidle")

    page.get_by_role("tab", name=SELECTORES["tab_ips"]).locator("span").click()

    _seleccionar_combobox(page, SELECTORES["tipo_doc_ips"], config["tipo_doc_ips"])
    page.get_by_role("textbox", name=SELECTORES["nit_ips"]).fill(config["nit_ips"])

    _seleccionar_combobox(page, SELECTORES["tipo_doc_usuario"], config["tipo_doc_usuario"])
    page.get_by_role("textbox", name=SELECTORES["id_usuario"]).fill(config["id_usuario"])

    page.get_by_role("textbox", name=SELECTORES["password"]).fill(password)
    page.get_by_role("button", name=SELECTORES["btn_ingresar"]).click()
    page.wait_for_load_state("networkidle")


def _ir_a_registro_dir(page) -> None:
    """Navega al módulo de Registro de Direccionamiento dentro de la sesión activa.

    Hace clic en el botón DIRECCIONAMIENTOS del menú principal y luego en
    la opción REGISTRAR DIRECCIONAMIENTO. Espera ``networkidle`` para asegurar
    que la página esté completamente cargada antes de continuar.

    Args:
        page: Instancia de ``playwright.sync_api.Page`` con sesión iniciada.
    """
    page.get_by_role("button", name=SELECTORES["btn_direccionamientos"], exact=True).click()
    el = page.get_by_text(SELECTORES["registrar_direccionamiento"])
    el.wait_for(state="visible")
    el.click()
    page.wait_for_load_state("networkidle")


def _seleccionar_sede(page, config: dict) -> None:
    """Selecciona la sucursal y sede IPS en el modal de configuración inicial.

    El modal aparece automáticamente al entrar al módulo de Direccionamiento.
    Usa ``_seleccionar_combobox`` con el contenedor del modal para evitar
    ambigüedades con otros comboboxes de la página.

    Args:
        page: Instancia de ``playwright.sync_api.Page``.
        config: Configuración del robot. Se usan ``config["navegacion"]["sucursal"]``
                y ``config["navegacion"]["sede"]``.
    """
    sucursal = config["navegacion"]["sucursal"]
    sede = config["navegacion"]["sede"]

    ventana = page.locator("kendo-window div").filter(has_text=SELECTORES["ventana_sede"])
    ventana.wait_for(state="visible")

    _seleccionar_combobox(page, SELECTORES["combobox_sucursal"], sucursal, contenedor=ventana)
    _seleccionar_combobox(page, SELECTORES["combobox_sede"], sede, contenedor=ventana)
    page.get_by_role("button", name=SELECTORES["btn_aceptar"]).click()


def _datos_vacio() -> dict:
    """Retorna un diccionario con los campos adicionales web en blanco.

    Se usa cuando el paciente no existe, el nombre no coincide, o el código
    de servicio no se encontró. Garantiza que las columnas H-L del Excel
    queden vacías en lugar de causar un ``KeyError``.

    Returns:
        Dict con claves ``paquete``, ``cantidad``, ``sede``,
        ``nombre_convenio`` y ``fecha_vencimiento``, todos con valor ``""``.
    """
    return {"paquete": "", "cantidad": "", "sede": "", "nombre_convenio": "", "fecha_vencimiento": ""}


def _extraer_datos_fila_web(page, cod_fact: str) -> dict:
    """Extrae los 5 campos adicionales de la fila que coincide con cod_fact en la tabla.

    Ejecuta JavaScript en el contexto de la página para navegar el DOM del
    kendo-grid sin depender de selectores CSS frágiles. Localiza el índice de
    cada columna de cabecera dinámicamente y luego recorre las filas hasta
    encontrar la que inicia con ``cod_fact``.

    Args:
        page: Instancia de ``playwright.sync_api.Page`` con la tabla visible.
        cod_fact: Código de facturación a buscar (búsqueda por ``startsWith``).

    Returns:
        Dict con claves ``paquete``, ``cantidad``, ``sede``,
        ``nombre_convenio`` y ``fecha_vencimiento`` con los valores
        encontrados, o ``_datos_vacio()`` si la fila no se encontró.
    """
    datos = page.evaluate(
        """(cod) => {
            const ths = Array.from(document.querySelectorAll('kendo-grid th'));
            const textos = ths.map(th => th.innerText.trim());

            const idx = {
                codigo:   textos.findIndex(t => t.includes('CÓDIGO SERVICIO')),
                clasif:   textos.findIndex(t => t.includes('CLASIFICACIÓN')),
                cantidad: textos.findIndex(t => t.includes('CANTIDAD SE')),
                sede:     textos.findIndex(t => t.includes('SEDE')),
                convenio: textos.findIndex(t => t.includes('NOMBRE CONVENIO')),
                fecha:    textos.findIndex(t => t.includes('FECHA VENCIMIENTO')),
            };

            const trs = Array.from(document.querySelectorAll(
                'kendo-grid-list table tbody tr:not(.k-grid-norecords)'
            ));

            for (const tr of trs) {
                const tds = tr.querySelectorAll('td');
                const codigo = tds[idx.codigo] ? tds[idx.codigo].textContent.trim() : '';
                if (codigo.startsWith(cod)) {
                    return {
                        paquete:           tds[idx.clasif]   ? tds[idx.clasif].textContent.trim()   : '',
                        cantidad:          tds[idx.cantidad] ? tds[idx.cantidad].textContent.trim() : '',
                        sede:              tds[idx.sede]     ? tds[idx.sede].textContent.trim()     : '',
                        nombre_convenio:   tds[idx.convenio] ? tds[idx.convenio].textContent.trim() : '',
                        fecha_vencimiento: tds[idx.fecha]    ? tds[idx.fecha].textContent.trim()    : '',
                    };
                }
            }
            return null;
        }""",
        cod_fact
    )
    return datos if datos else _datos_vacio()


def _obtener_codigos_web(page) -> list[str]:
    """Extrae todos los códigos de servicio del kendo-grid navegando por todas las páginas.

    Itera hasta 20 páginas haciendo clic en el botón "Siguiente" del paginador
    Kendo. En cada página extrae los valores de la columna ``CÓDIGO SERVICIO``
    via JavaScript. Retorna la unión de todos como set (sin duplicados).

    Args:
        page: Instancia de ``playwright.sync_api.Page`` con el grid visible
              y filtrado por número de autorización.

    Returns:
        Lista de strings con todos los códigos de servicio encontrados
        en la autorización consultada.
    """
    codigos = []
    limite_paginas = 0

    while limite_paginas < 20:
        codigos_pagina = page.evaluate('''() => {
            const ths = Array.from(document.querySelectorAll('kendo-grid th'));
            const colIndex = ths.findIndex(th => th.innerText.includes('CÓDIGO SERVICIO'));
            if (colIndex === -1) return [];
            const trs = Array.from(document.querySelectorAll(
                'kendo-grid-list table tbody tr:not(.k-grid-norecords)'
            ));
            return trs.map(tr => {
                const tds = tr.querySelectorAll('td');
                return tds[colIndex] ? tds[colIndex].textContent.trim() : '';
            }).filter(c => c !== '');
        }''')

        codigos.extend(codigos_pagina)

        avanzo = page.evaluate('''() => {
            const pager = document.querySelector(
                "#k-tabstrip-tabpanel-1 > form > div.row.ng-star-inserted > kendo-grid > kendo-pager"
            ) || document;
            const nextGroup = pager.querySelector('kendo-pager-next-buttons');
            const nextBtn = nextGroup
                ? nextGroup.querySelector('button')
                : pager.querySelector('button[title*="next" i], .k-pager-nav:not(.k-pager-last)');
            if (nextBtn && !nextBtn.disabled && !nextBtn.classList.contains('k-disabled')) {
                nextBtn.click();
                return true;
            }
            return false;
        }''')

        if avanzo:
            page.wait_for_timeout(2000)
            limite_paginas += 1
        else:
            break

    return list(set(codigos))


def _procesar_fila(page, fila: dict, config: dict, logger, idx: int, total: int) -> dict:
    """Consulta un paciente en la plataforma web y retorna el estado de verificación.

    Flujo completo por fila:
    1. Ingresa el ID del paciente y presiona CONSULTAR.
    2. Detecta el popup "No se encontró el afiliado" (timeout 3000 ms) y lo
       cierra retornando ``ESTADO_PAC_NO_EXISTE`` si aparece.
    3. Cierra el alert de PAC PLAN ALFA si está presente.
    4. Valida que los primeros 5 caracteres del nombre web coincidan con el Excel.
    5. Navega a CONSULTAR DIRECCIONAMIENTO y filtra por número de autorización.
    6. Si no hay resultados, aplica el "truco de fechas": regresa 6 meses y
       vuelve a consultar (maneja autorizaciones antiguas que cambian de año).
    7. Extrae los códigos del grid con paginación y busca ``cod_fact``.
    8. Si encuentra coincidencia, extrae los 5 campos adicionales.

    Args:
        page: Instancia de ``playwright.sync_api.Page`` posicionada en el
              módulo de REGISTRAR DIRECCIONAMIENTO.
        fila: Dict con las claves ``id_paciente``, ``nombre_paciente``,
              ``autorizacion`` y ``cod_fact`` leídas del Excel.
        config: Configuración del robot.
        logger: Logger centralizado.
        idx: Número de fila actual (para mensajes de log).
        total: Total de filas pendientes (para mensajes de log).

    Returns:
        Dict con la clave ``"estado"`` (uno de los valores ``ESTADO_*``) y
        las claves ``paquete``, ``cantidad``, ``sede``, ``nombre_convenio``,
        ``fecha_vencimiento``. Si no se encuentra el registro, estos últimos
        5 campos retornan vacíos.
    """
    id_pac  = str(fila["id_paciente"]).strip()
    nombre  = str(fila["nombre_paciente"]).strip().upper()
    aut     = str(fila["autorizacion"]).strip()
    cod     = str(fila["cod_fact"]).strip()

    logger.info("[Fila %d/%d] Consultando paciente ID=%s | AUT=%s", idx, total, id_pac, aut)

    # Ingresar cédula y consultar
    campo_cedula = page.get_by_role("textbox", name=SELECTORES["campo_id_paciente"])
    campo_cedula.wait_for(state="visible")
    campo_cedula.click()
    campo_cedula.press("Control+a")
    campo_cedula.press("Backspace")
    campo_cedula.type(id_pac, delay=30)
    # Verificar que el valor quedó escrito antes de consultar
    if campo_cedula.input_value().strip() != id_pac:
        logger.warning("[Fila %d/%d] Reintentando escritura de cédula ID=%s", idx, total, id_pac)
        campo_cedula.click()
        campo_cedula.press("Control+a")
        campo_cedula.type(id_pac, delay=60)
    page.get_by_role("button", name=SELECTORES["btn_consultar"]).first.click()

    # Manejar popup "No se encontró el afiliado"
    try:
        dialogo_no_afiliado = page.locator("kendo-dialog").filter(
            has_text=re.compile(r"No se encontró el afiliado", re.IGNORECASE)
        )
        dialogo_no_afiliado.wait_for(state="visible", timeout=3000)
        dialogo_no_afiliado.get_by_role("button", name="Aceptar").click()
        logger.warning("[Fila %d/%d] Afiliado no encontrado en sistema: ID=%s", idx, total, id_pac)
        return {"estado": ESTADO_PAC_NO_EXISTE, **_datos_vacio()}
    except PlaywrightTimeout:
        pass  # No apareció → el afiliado existe, continuar

    # Manejar alerta PAC PLAN ALFA
    try:
        dialogo = page.locator("app-info-afiliado app-ventana kendo-dialog").filter(
            has_text=re.compile(r"AFILIADO PAC PLAN ALFA", re.IGNORECASE)
        )
        dialogo.wait_for(state="visible", timeout=2500)
        dialogo.locator("kendo-dialog-actions button").click()
    except PlaywrightTimeout:
        pass

    page.wait_for_timeout(1500)

    # Validar nombre (primeros 5 caracteres)
    nombre_web = page.get_by_role(
        "textbox", name=SELECTORES["campo_nombre"]
    ).input_value().strip().upper()

    if not nombre_web:
        logger.warning("[Fila %d/%d] Paciente no encontrado: ID=%s", idx, total, id_pac)
        return {"estado": ESTADO_PAC_NO_EXISTE, **_datos_vacio()}

    if nombre_web[:5] != nombre[:5]:
        logger.warning("[Fila %d/%d] Nombre no coincide. Web='%s' | Excel='%s'",
                       idx, total, nombre_web, nombre)
        return {"estado": ESTADO_NOMBRE_NO_OK, **_datos_vacio()}

    # Consultar direccionamientos
    page.get_by_text(SELECTORES["consultar_dir"]).click()
    page.get_by_role("button", name=SELECTORES["btn_consultar"]).last.click()

    # Truco de fechas si no hay resultados
    try:
        alerta = page.locator("div").filter(
            has_text=re.compile(r"^No se encontraron coincidencias en la consulta$")
        )
        alerta.wait_for(state="visible", timeout=3000)
        page.get_by_role("button", name="Aceptar").click()

        import pandas as pd
        fecha1 = page.get_by_role("spinbutton", name="null").first
        fecha2 = page.get_by_role("spinbutton", name="null").nth(1)
        v1, v2 = fecha1.input_value(), fecha2.input_value()

        if v1 and v2:
            d1 = pd.to_datetime(v1, dayfirst=True) - pd.DateOffset(months=6)
            d2 = pd.to_datetime(v2, dayfirst=True) - pd.DateOffset(months=6)

            for campo, valor in [(fecha1, d1), (fecha2, d2)]:
                campo.click()
                page.keyboard.press("Control+a")
                page.keyboard.press("Backspace")
                campo.type(valor.strftime("%d/%m/%Y"), delay=50)

            page.keyboard.press("Enter")
            page.get_by_role("button", name=SELECTORES["btn_consultar"]).last.click()

            try:
                page.locator("kendo-grid .k-loading-mask").wait_for(state="visible", timeout=1000)
            except PlaywrightTimeout:
                pass
            try:
                page.locator("kendo-grid .k-loading-mask").wait_for(state="hidden", timeout=8000)
            except PlaywrightTimeout:
                pass
            page.wait_for_timeout(1000)

    except PlaywrightTimeout:
        pass  # Resultados encontrados sin truco de fechas

    # Filtrar por autorización
    filtro = page.get_by_role("textbox", name=SELECTORES["filtro_aut"], exact=True)
    filtro.wait_for(state="visible")
    filtro.click()
    page.keyboard.press("Control+a")
    page.keyboard.press("Backspace")
    page.wait_for_timeout(2500)
    filtro.fill(aut)
    page.wait_for_timeout(3500)

    # Extraer códigos de la tabla
    codigos_web = _obtener_codigos_web(page)
    logger.info("[Fila %d/%d] Códigos en web para AUT=%s: %s", idx, total, aut, codigos_web)

    # Validar si cod_fact está en los resultados y extraer datos adicionales
    if any(str(cw).startswith(cod) for cw in codigos_web):
        datos_adicionales = _extraer_datos_fila_web(page, cod)
        logger.info("[Fila %d/%d] ENCONTRADO: cod=%s | Paquete=%s | Convenio=%s | Vence=%s",
                    idx, total, cod,
                    datos_adicionales.get("paquete", ""),
                    datos_adicionales.get("nombre_convenio", ""),
                    datos_adicionales.get("fecha_vencimiento", ""))
        return {"estado": ESTADO_ENCONTRADO, **datos_adicionales}
    else:
        logger.warning("[Fila %d/%d] NO ENCONTRADO: cod_fact=%s en AUT=%s", idx, total, cod, aut)
        return {"estado": ESTADO_NO_ENCONTRADO, **_datos_vacio()}


def ejecutar(config: dict, logger, datos_entrada: Any = None) -> Tuple[bool, Any]:
    """Verifica en la web de SaludTotal todas las filas pendientes del consolidado.

    Abre una sesión de Playwright Chromium, hace login, navega al módulo de
    Direccionamiento y procesa secuencialmente cada fila con ``ESTADO_ROBOT``
    vacío. Escribe el resultado en el Excel en tiempo real (fila a fila) para
    no perder progreso ante interrupciones. Reutiliza la sesión web mientras el
    paciente no cambie; si cambia, navega de vuelta al formulario sin cerrar
    el browser. Ante fallo por intento aplica hasta ``reintentos_max`` intentos
    reiniciando la sesión web completa.

    Args:
        config: Configuración del robot. Se usan las claves
                ``config["archivos"]["consolidado"]``,
                ``config["parametros"]["reintentos_max"]``,
                ``config["parametros"]["guardar_cada_n_registros"]``,
                ``config["tiempos"]["timeout_general"]`` y todas las de
                ``config["navegacion"]``.
        logger: Logger centralizado.
        datos_entrada: No utilizado. Existe por convención de arquitectura.

    Returns:
        Tupla ``(True, resumen)`` donde ``resumen`` es::

            {"verificados": int, "pendientes": int}

        Retorna ``(True, {"verificados": 0, "pendientes": 0})`` si
        ``consolidado.xlsx`` no existe o no tiene filas pendientes.
        Retorna ``(False, None)`` ante error crítico de Playwright.
    """
    nombre_proceso = "p05_navegacion_web"
    logger.info(">> INICIO subproceso: %s", nombre_proceso)

    ruta_consolidado = config["archivos"]["consolidado"]
    reintentos_max   = config["parametros"]["reintentos_max"]
    guardar_cada     = config["parametros"]["guardar_cada_n_registros"]
    tiempos          = config.get("tiempos", {})

    if not os.path.exists(ruta_consolidado):
        logger.info("consolidado.xlsx no existe aún. No hay filas que verificar.")
        logger.info("<< FIN subproceso: %s — OK (sin datos)", nombre_proceso)
        return True, {"verificados": 0, "pendientes": 0}

    # Cargar Excel y detectar filas pendientes (ESTADO_ROBOT vacío)
    wb = openpyxl.load_workbook(ruta_consolidado)
    ws = wb.active

    filas_pendientes = []
    for row_idx in range(2, ws.max_row + 1):
        estado = ws.cell(row=row_idx, column=COL_ESTADO_ROBOT).value
        if not estado or str(estado).strip() == "":
            filas_pendientes.append({
                "row_idx":        row_idx,
                "id_paciente":    str(ws.cell(row=row_idx, column=COL_ID_PACIENTE).value or "").strip(),
                "nombre_paciente":str(ws.cell(row=row_idx, column=COL_NOMBRE).value or "").strip(),
                "autorizacion":   str(ws.cell(row=row_idx, column=COL_AUTORIZACION).value or "").strip(),
                "cod_fact":       str(ws.cell(row=row_idx, column=COL_COD_FACT).value or "").strip(),
            })

    if not filas_pendientes:
        logger.info("No hay filas pendientes de verificar en consolidado.xlsx")
        logger.info("<< FIN subproceso: %s — OK", nombre_proceso)
        return True, {"verificados": 0, "pendientes": 0}

    total = len(filas_pendientes)
    logger.info("Filas pendientes de verificación web: %d", total)

    pw = None
    browser = None

    try:
        pw = sync_playwright().start()
        browser = pw.chromium.launch(
            headless=config["navegacion"]["headless"],
            slow_mo=config["navegacion"]["slow_mo"],
            args=["--start-maximized"]
        )
        context = browser.new_context(no_viewport=True)
        page = context.new_page()
        page.set_default_timeout(tiempos.get("timeout_general", 120) * 1000)

        logger.info("Iniciando login en SaludTotal...")
        _login(page, config)
        logger.info("Login exitoso.")

        logger.info("Navegando a REGISTRAR DIRECCIONAMIENTO...")
        _ir_a_registro_dir(page)
        _seleccionar_sede(page, config)
        logger.info("Sede configurada: %s", config["navegacion"]["sede"])

        verificados = 0
        paciente_actual = None

        for contador, fila in enumerate(filas_pendientes, start=1):
            resultado = {"estado": ESTADO_ERROR, **_datos_vacio()}
            id_pac = fila["id_paciente"]

            # Validar que los 4 campos requeridos estén presentes
            campos_vacios = [c for c in ("id_paciente", "nombre_paciente", "autorizacion", "cod_fact")
                             if not fila.get(c)]
            if campos_vacios:
                logger.warning(
                    "[Fila %d/%d] Campos vacíos en Excel (%s) — se marca 'Falta información' y se omite",
                    contador, total, ", ".join(campos_vacios)
                )
                ws.cell(row=fila["row_idx"], column=COL_ESTADO_ROBOT, value="Falta información")
                wb.save(ruta_consolidado)
                verificados += 1
                continue

            for intento in range(1, reintentos_max + 1):
                try:
                    # Si cambia el paciente, navegar de vuelta al formulario
                    if paciente_actual is not None and id_pac != paciente_actual:
                        logger.info("[Fila %d/%d] Nuevo paciente — volviendo al formulario",
                                    contador, total)
                        el = page.get_by_text(SELECTORES["registrar_direccionamiento"])
                        el.wait_for(state="visible")
                        el.click()
                        page.wait_for_load_state("networkidle")

                    resultado = _procesar_fila(page, fila, config, logger, contador, total)
                    paciente_actual = id_pac
                    break

                except Exception as e:
                    logger.warning("[Fila %d/%d] Intento %d/%d fallido: %s",
                                   contador, total, intento, reintentos_max, str(e))
                    if intento < reintentos_max:
                        # Reiniciar sesión web
                        try:
                            logger.info("Reiniciando sesión web...")
                            page.reload()
                            page.wait_for_load_state("networkidle")
                            _login(page, config)
                            _ir_a_registro_dir(page)
                            _seleccionar_sede(page, config)
                            paciente_actual = None
                        except Exception as restart_err:
                            logger.error("Error al reiniciar sesión: %s", str(restart_err))

            # Escribir estado + datos adicionales en Excel en tiempo real
            ws.cell(row=fila["row_idx"], column=COL_ESTADO_ROBOT,      value=resultado["estado"])
            ws.cell(row=fila["row_idx"], column=COL_PAQUETE,           value=resultado.get("paquete", ""))
            ws.cell(row=fila["row_idx"], column=COL_CANTIDAD,          value=resultado.get("cantidad", ""))
            ws.cell(row=fila["row_idx"], column=COL_SEDE,              value=resultado.get("sede", ""))
            ws.cell(row=fila["row_idx"], column=COL_NOMBRE_CONVENIO,   value=resultado.get("nombre_convenio", ""))
            ws.cell(row=fila["row_idx"], column=COL_FECHA_VENCIMIENTO, value=resultado.get("fecha_vencimiento", ""))
            wb.save(ruta_consolidado)
            verificados += 1
            logger.info("[Fila %d/%d] ESTADO_ROBOT: %s", contador, total, resultado["estado"])

            if contador % guardar_cada == 0:
                logger.info("Guardado parcial en fila %d/%d", contador, total)

        logger.info("Verificación web completada — %d/%d filas procesadas", verificados, total)
        logger.info("<< FIN subproceso: %s — OK", nombre_proceso)
        return True, {"verificados": verificados, "pendientes": total - verificados}

    except Exception as e:
        logger.error("<< FIN subproceso: %s — ERROR: %s", nombre_proceso, str(e))
        return False, None

    finally:
        if browser:
            browser.close()
        if pw:
            pw.stop()
