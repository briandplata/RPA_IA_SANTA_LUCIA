"""
Módulo: processes.p03_excel
Descripción: Tercer subproceso del pipeline. Verifica duplicados por número de
             autorización y escribe los resultados en dos archivos Excel:
             - ``pacientes/procesados/consolidado.xlsx`` (exitosos únicos, 12 columnas)
             - ``pacientes/no_procesados/errores.xlsx`` (OCR fallido o duplicados)
             Columna A: fecha/hora de lectura. Columna G: ESTADO_ROBOT vacío = pendiente.
             Las columnas H-L las completa p05 en tiempo real tras la verificación web.
Autor: Brayand Javier Gomez Plata
"""

import os
from datetime import datetime
from typing import Any, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


CABECERAS_CONSOLIDADO = [
    "Fecha_Hora", "Archivo", "ID_Paciente",
    "Nombre_Paciente", "Autorizacion", "Cod_Fact", "ESTADO_ROBOT",
    "Paquete", "Cantidad", "Sede", "Nombre_Convenio", "Fecha_Vencimiento"
]

CABECERAS_ERRORES = [
    "Fecha_Hora", "Archivo", "Motivo_Error", "Detalle"
]

# Índices de columna en consolidado (0-based)
COL_AUTORIZACION = 4   # Columna E
COL_ESTADO_ROBOT = 6   # Columna G


def _cargar_o_crear_excel(ruta: str, cabeceras: list[str]) -> openpyxl.Workbook:
    """Carga el Excel existente o crea uno nuevo con cabeceras estilizadas.

    Si el archivo no existe lo crea con la fila de cabecera formateada:
    fondo azul oscuro (#1F4E79), texto blanco en negrita y alineación centrada.
    El directorio padre se crea automáticamente si no existe.

    Args:
        ruta: Ruta al archivo ``.xlsx``.
        cabeceras: Lista de nombres de columna en el orden correcto.

    Returns:
        ``openpyxl.Workbook`` listo para agregar filas.
    """
    if os.path.exists(ruta):
        return openpyxl.load_workbook(ruta)

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, cabecera in enumerate(cabeceras, start=1):
        cell = ws.cell(row=1, column=col_idx, value=cabecera)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment

    ws.row_dimensions[1].height = 20
    os.makedirs(os.path.dirname(ruta), exist_ok=True)
    return wb


def _obtener_autorizaciones_existentes(ruta: str) -> set[str]:
    """Lee consolidado.xlsx y retorna el conjunto de números AUT ya registrados.

    Permite detectar duplicados antes de insertar nuevas filas. Lee solo la
    columna E (índice 4, ``COL_AUTORIZACION``) para minimizar memoria.

    Args:
        ruta: Ruta al archivo ``consolidado.xlsx``.

    Returns:
        ``set`` de strings con cada número de autorización registrado,
        o ``set`` vacío si el archivo no existe todavía.
    """
    if not os.path.exists(ruta):
        return set()

    wb = openpyxl.load_workbook(ruta)
    ws = wb.active
    autorizaciones = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[COL_AUTORIZACION]:
            autorizaciones.add(str(row[COL_AUTORIZACION]).strip())

    return autorizaciones


def ejecutar(config: dict, logger, datos_entrada: Any = None) -> Tuple[bool, Any]:
    """Escribe los resultados del OCR en los archivos Excel de consolidado y errores.

    Flujo:
    1. Lee las autorizaciones ya existentes en ``consolidado.xlsx`` para
       detectar duplicados sin releer el archivo completo en cada iteración.
    2. Registra los errores de OCR en ``errores.xlsx``.
    3. Para cada exitoso: si el AUT es duplicado lo envía a ``errores.xlsx``
       y marca ``item["_duplicado"] = True``; si es nuevo lo agrega al
       consolidado con ESTADO_ROBOT vacío (columna G) y ``_duplicado = False``.
    4. Guarda ambos archivos al finalizar.

    Args:
        config: Configuración de ``config.yaml``. Se usan las claves
                ``config["archivos"]["consolidado"]`` y
                ``config["archivos"]["errores"]``.
        logger: Logger centralizado.
        datos_entrada: Dict ``{"exitosos": [...], "errores": [...]}``
                       tal como lo retorna ``p02_ocr_ia.ejecutar()``.

    Returns:
        Tupla ``(True, datos_entrada)`` donde cada item en ``datos_entrada``
        ha sido anotado con la clave ``"_duplicado": bool``.
        Retorna ``(False, None)`` ante errores críticos.

    Note:
        ``datos_entrada`` se retorna modificado en lugar de una copia,
        para que ``p04_mover_archivos`` pueda leer el flag ``_duplicado``
        sin trabajo adicional.
    """
    nombre_proceso = "p03_excel"
    logger.info(">> INICIO subproceso: %s", nombre_proceso)

    if not datos_entrada:
        logger.error("No se recibieron datos del proceso OCR")
        return False, None

    ruta_consolidado = config["archivos"]["consolidado"]
    ruta_errores = config["archivos"]["errores"]

    exitosos: list[dict] = datos_entrada.get("exitosos", [])
    errores: list[dict] = datos_entrada.get("errores", [])

    try:
        autorizaciones_existentes = _obtener_autorizaciones_existentes(ruta_consolidado)
        logger.info("Autorizaciones ya registradas en consolidado: %d", len(autorizaciones_existentes))

        wb_consolidado = _cargar_o_crear_excel(ruta_consolidado, CABECERAS_CONSOLIDADO)
        wb_errores = _cargar_o_crear_excel(ruta_errores, CABECERAS_ERRORES)
        ws_consolidado = wb_consolidado.active
        ws_errores = wb_errores.active

        nuevos_consolidado = 0
        nuevos_errores = len(errores)

        # Escribir errores de OCR
        for item in errores:
            ws_errores.append([
                item["fecha_hora"],
                item["archivo"],
                item["motivo"],
                item["detalle"]
            ])
            logger.info("Registrado en errores.xlsx: %s — %s", item["archivo"], item["motivo"])

        # Procesar exitosos verificando duplicados
        for item in exitosos:
            aut = str(item["autorizacion"]).strip()

            if aut in autorizaciones_existentes:
                fecha_hora = datetime.now().strftime("%d/%m/%Y %I:%M:%S %p").lower()
                ws_errores.append([
                    fecha_hora,
                    item["archivo"],
                    "DUPLICADO",
                    f"La autorización {aut} ya existe en consolidado.xlsx"
                ])
                nuevos_errores += 1
                logger.warning("DUPLICADO detectado: AUT=%s | Archivo=%s", aut, item["archivo"])
                item["_duplicado"] = True
            else:
                ws_consolidado.append([
                    item["fecha_hora"],     # A: Fecha_Hora
                    item["archivo"],        # B: Archivo
                    item["id_paciente"],    # C: ID_Paciente
                    item["nombre_paciente"],# D: Nombre_Paciente
                    item["autorizacion"],   # E: Autorizacion
                    item["cod_fact"],       # F: Cod_Fact
                    "",                     # G: ESTADO_ROBOT (vacío = pendiente)
                    "", "", "", "", ""      # H-L: datos web (se llenan en p05)
                ])
                autorizaciones_existentes.add(aut)
                nuevos_consolidado += 1
                item["_duplicado"] = False
                logger.info("Registrado en consolidado.xlsx: AUT=%s | %s",
                            aut, item["nombre_paciente"])

        wb_consolidado.save(ruta_consolidado)
        wb_errores.save(ruta_errores)

        logger.info("Excel guardados — Consolidado: +%d registros | Errores: +%d registros",
                    nuevos_consolidado, nuevos_errores)
        logger.info("<< FIN subproceso: %s — OK", nombre_proceso)
        return True, datos_entrada

    except Exception as e:
        logger.error("<< FIN subproceso: %s — ERROR: %s", nombre_proceso, str(e))
        return False, None
